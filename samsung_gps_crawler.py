"""
삼성 멤버스 커뮤니티 GPS 게시글 크롤러
--------------------------------------
사용법:
  1. 패키지 설치:
     pip install selenium webdriver-manager openpyxl

  2. 실행:
     python samsung_gps_crawler.py

결과:  samsung_gps_posts.xlsx 파일이 생성됩니다.
"""

import time
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── 설정 ──────────────────────────────────────────────
SEARCH_URL = (
    "https://r1.community.samsung.com/t5/forums/searchpage/tab/message"
    "?advanced=false&allow_punctuation=false"
    "&filter=location&location=category:kr-community&q=gps"
)
MAX_PAGES   = 10          # 크롤링할 최대 페이지 수 (None = 전체)
OUTPUT_FILE = "samsung_gps_posts.xlsx"
WAIT_SEC    = 3           # 페이지 로드 대기 시간(초)
# ──────────────────────────────────────────────────────


def build_driver():
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)


def safe_text(el, css):
    try:
        return el.find_element(By.CSS_SELECTOR, css).text.strip()
    except Exception:
        return ""


def safe_attr(el, css, attr):
    try:
        return el.find_element(By.CSS_SELECTOR, css).get_attribute(attr) or ""
    except Exception:
        return ""


def crawl_posts(driver):
    posts = []
    page  = 1

    driver.get(SEARCH_URL)
    time.sleep(WAIT_SEC + 1)

    while True:
        print(f"  페이지 {page} 크롤링 중...")
        try:
            WebDriverWait(driver, 12).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, "li.search-result"))
            )
        except Exception:
            print("  → 게시글을 찾을 수 없습니다. 종료합니다.")
            break

        items = driver.find_elements(By.CSS_SELECTOR, "li.search-result")
        if not items:
            break

        for item in items:
            title   = safe_text(item, "h3.search-result-title a, .lia-link-navigation")
            link    = safe_attr(item, "h3.search-result-title a, .lia-link-navigation", "href")
            author  = safe_text(item, ".UserName a, .lia-user-name a")
            date    = safe_text(item, "span.local-date, .search-result-date, time")
            board   = safe_text(item, ".search-result-label-board a, .search-snippet-board a")
            preview = safe_text(item, ".search-result-content p, .search-snippet")
            kudos   = safe_text(item, ".search-result-kudos, .kudos-count")
            replies = safe_text(item, ".search-result-replies, .reply-count")

            if not title:
                continue

            posts.append({
                "번호":      len(posts) + 1,
                "제목":      title,
                "작성자":    author,
                "작성일":    date,
                "게시판":    board,
                "내용 요약": preview,
                "좋아요":    kudos,
                "댓글수":    replies,
                "URL":       link,
            })

        if MAX_PAGES and page >= MAX_PAGES:
            break

        try:
            next_btn = driver.find_element(
                By.CSS_SELECTOR,
                "a.lia-link-ticket-post-action[rel='next'], "
                "li.lia-paging-page-next a, "
                ".pagination-next a"
            )
            next_btn.click()
            page += 1
            time.sleep(WAIT_SEC)
        except Exception:
            print("  → 다음 페이지 없음.")
            break

    return posts


def save_excel(posts, filepath):
    wb = Workbook()
    ws = wb.active
    ws.title = "GPS 게시글"

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color="1E4D8C")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    alt_fill = PatternFill("solid", start_color="EEF3FB")

    headers    = ["번호", "제목", "작성자", "작성일", "게시판", "내용 요약", "좋아요", "댓글수", "URL"]
    col_widths = [6,      45,    15,      18,      20,      50,        8,      8,      60]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22

    for r_idx, post in enumerate(posts, start=2):
        alt = (r_idx % 2 == 0)
        for c_idx, key in enumerate(headers, start=1):
            val  = post.get(key, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            cell.font   = Font(name="Arial", size=10)
            if alt:
                cell.fill = alt_fill
            if key == "URL" and val.startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=10, color="1155CC", underline="single")
            cell.alignment = (
                center_align if key in ("번호", "작성일", "좋아요", "댓글수") else left_align
            )

    ws_sum = wb.create_sheet("수집 정보")
    ws_sum["A1"] = "수집 일시";  ws_sum["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_sum["A2"] = "검색 URL";   ws_sum["B2"] = SEARCH_URL
    ws_sum["A3"] = "총 게시글 수"; ws_sum["B3"] = len(posts)
    for cell in [ws_sum["A1"], ws_sum["A2"], ws_sum["A3"]]:
        cell.font = Font(name="Arial", bold=True)
    ws_sum.column_dimensions["A"].width = 18
    ws_sum.column_dimensions["B"].width = 60

    wb.save(filepath)
    print(f"\n✅ 저장 완료: {filepath}  (총 {len(posts)}건)")


def main():
    print("=" * 50)
    print("삼성 멤버스 GPS 게시글 크롤러")
    print("=" * 50)
    driver = build_driver()
    try:
        posts = crawl_posts(driver)
        if posts:
            save_excel(posts, OUTPUT_FILE)
        else:
            print("수집된 게시글이 없습니다.")
    finally:
        driver.quit()


if __name__ == "__main__":
    main()

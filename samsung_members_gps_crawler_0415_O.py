"""
삼성 멤버스 커뮤니티 GPS 검색 결과 크롤러
==========================================
대상 URL:
  https://r1.community.samsung.com/t5/forums/searchpage/tab/message
  ?advanced=false&allow_punctuation=false
  &filter=location&location=category:kr-community&q=gps

특징:
  - Khoros 플랫폼 기반 SPA → Selenium으로 JS 렌더링 후 파싱
  - 게시글 목록 페이지에서 제목·작성자·날짜·좋아요·댓글수·URL 수집
  - 각 게시글 상세 페이지로 이동해 본문 전체 수집
  - 작성일 기준 최근 2주 이내 게시글만 저장
  - 결과를 openpyxl로 엑셀 저장

실행 방법:
  pip install selenium webdriver-manager openpyxl
  python samsung_members_crawler.py
"""

import time
import re
from datetime import datetime, timedelta, timezone

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 설정
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SEARCH_URL = (
    "https://r1.community.samsung.com/t5/forums/searchpage/tab/message"
    "?advanced=false&allow_punctuation=false"
    "&filter=location&location=category:kr-community&q=gps"
)
OUTPUT_FILE  = "samsung_gps_2weeks.xlsx"
DAYS_LIMIT   = 14          # 최근 2주
PAGE_TIMEOUT = 15          # 페이지 로드 대기(초)
DETAIL_WAIT  = 4           # 상세 페이지 로드 대기(초)
MAX_PAGES    = 20          # 최대 탐색 페이지 수 (None = 무제한)
HEADLESS     = True        # False 로 바꾸면 브라우저 창이 뜸


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Khoros 플랫폼 CSS 셀렉터 (삼성멤버스 기준)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# [목록 페이지] 검색결과 행
SEL_RESULT_ITEMS = "li.search-result, li[class*='search-result']"

# 목록 행 내부 셀렉터
SEL_TITLE   = (
    "h3.search-result-title a, "
    ".lia-search-results-message-body a[href*='/t5/'], "
    "a[class*='lia-link-navigation'][href*='/t5/']"
)
SEL_AUTHOR  = ".UserName a, .lia-user-name-link, a[class*='UserName']"
SEL_DATE_LIST = (
    "span.local-date, "
    "span[class*='search-result-date'], "
    "time[class*='local-date'], "
    ".lia-message-stats span.local-date"
)
SEL_KUDOS_LIST = (
    ".search-result-kudos, "
    ".lia-kudos-count, "
    "span[class*='kudos']"
)
SEL_REPLIES_LIST = (
    ".search-result-replies, "
    ".lia-message-replies, "
    "span[class*='replies']"
)
SEL_BOARD   = (
    ".search-result-label-board a, "
    ".lia-quilt-column-alley-left a[class*='PageBreadcrumb'], "
    "span[class*='search-result-label-board'] a"
)
SEL_NEXT_PAGE = (
    "a.lia-link-ticket-post-action[rel='next'], "
    "li.lia-paging-page-next a, "
    "a[class*='page-link'][rel='next']"
)

# [상세 페이지] 본문·날짜·좋아요 셀렉터
# Khoros 상세 페이지 구조:
#   본문:    div[class*='lia-message-body-content']
#   날짜:    span.local-date  또는  time.local-date (첫 번째)
#   좋아요:  span[class*='lia-component-kudos-widget-button-count']
#             또는 .kudos-count  또는  span.lia-button-image-kudos-count
SEL_BODY    = (
    "div.lia-message-body-content, "
    "div[class*='lia-message-body-content'], "
    "div[id*='bodyDisplay']"
)
SEL_DATE_DETAIL = (
    "span.local-date, "
    "time.local-date, "
    "span[class*='local-date']"
)
SEL_KUDOS_DETAIL = (
    "span.lia-button-image-kudos-count, "
    "span[class*='lia-component-kudos-widget-button-count'], "
    "span[class*='kudos-count'], "
    ".kudos-count, "
    "span[class*='count'][class*='kudos']"
)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 날짜 파싱
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
KST = timezone(timedelta(hours=9))
TWO_WEEKS_AGO = datetime.now(KST) - timedelta(days=DAYS_LIMIT)

DATE_PATTERNS = [
    # 2024-08-22 02:13 PM  →  24시간 변환 후 파싱
    (r'(\d{2}-\d{2}-\d{4})\s+(\d{1,2}:\d{2})\s+(AM|PM)',
     lambda m: datetime.strptime(
         f"{m.group(1)} {m.group(2)} {m.group(3)}", "%m-%d-%Y %I:%M %p"
     ).replace(tzinfo=KST)),
    # ‎04-08-2026 05:01 PM
    (r'(\d{2}-\d{2}-\d{4})',
     lambda m: datetime.strptime(m.group(1), "%m-%d-%Y").replace(tzinfo=KST)),
    # 2026-04-08
    (r'(\d{4}-\d{2}-\d{2})',
     lambda m: datetime.strptime(m.group(1), "%Y-%m-%d").replace(tzinfo=KST)),
    # April 8, 2026 / Apr 8, 2026
    (r'([A-Za-z]{3,9})\s+(\d{1,2}),?\s+(\d{4})',
     lambda m: datetime.strptime(
         f"{m.group(1)} {m.group(2)} {m.group(3)}", "%B %d %Y"
     ).replace(tzinfo=KST)),
]

def parse_date(text: str):
    """날짜 문자열 → datetime (KST). 파싱 실패 시 None 반환."""
    text = text.strip().replace('\u200e', '').replace('\u200f', '')
    for pattern, converter in DATE_PATTERNS:
        m = re.search(pattern, text)
        if m:
            try:
                return converter(m)
            except ValueError:
                continue
    return None

def is_within_2weeks(dt) -> bool:
    if dt is None:
        return True   # 날짜 불명이면 일단 포함
    return dt >= TWO_WEEKS_AGO


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Selenium 드라이버 설정
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def build_driver():
    opts = Options()
    if HEADLESS:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
    svc = Service(ChromeDriverManager().install())
    drv = webdriver.Chrome(service=svc, options=opts)
    drv.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"}
    )
    return drv


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 헬퍼: 안전한 텍스트 / 속성 추출
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def safe_text(parent, css: str, default="") -> str:
    try:
        el = parent.find_element(By.CSS_SELECTOR, css)
        return (el.get_attribute("textContent") or el.text or "").strip()
    except NoSuchElementException:
        return default

def safe_attr(parent, css: str, attr: str, default="") -> str:
    try:
        el = parent.find_element(By.CSS_SELECTOR, css)
        return (el.get_attribute(attr) or "").strip()
    except NoSuchElementException:
        return default

def wait_for(driver, css: str, timeout=PAGE_TIMEOUT):
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_all_elements_located((By.CSS_SELECTOR, css))
    )


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 상세 페이지 → 본문 전체 + 날짜 + 좋아요 재수집
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def fetch_detail(driver, url: str) -> dict:
    """게시글 상세 페이지에서 본문 전체 / 정확한 날짜 / 좋아요 추출"""
    result = {"body": "", "date_str": "", "date": None, "kudos": "0"}
    try:
        driver.get(url)
        time.sleep(DETAIL_WAIT)

        # ── 본문 전체 ──
        try:
            body_el = driver.find_element(By.CSS_SELECTOR, SEL_BODY)
            result["body"] = (body_el.get_attribute("innerText") or body_el.text or "").strip()
        except NoSuchElementException:
            pass

        # ── 날짜 (상세 페이지 첫 번째 local-date) ──
        try:
            date_els = driver.find_elements(By.CSS_SELECTOR, SEL_DATE_DETAIL)
            for el in date_els:
                raw = (el.get_attribute("textContent") or el.text or "").strip()
                # data-friendly-date 속성도 시도
                friendly = el.get_attribute("data-friendly-date") or ""
                dt_text = friendly if friendly else raw
                dt = parse_date(dt_text)
                if dt:
                    result["date"] = dt
                    result["date_str"] = dt.strftime("%Y-%m-%d %H:%M")
                    break
        except Exception:
            pass

        # ── 좋아요 수 ──
        try:
            kudos_els = driver.find_elements(By.CSS_SELECTOR, SEL_KUDOS_DETAIL)
            for el in kudos_els:
                txt = (el.get_attribute("textContent") or el.text or "").strip()
                if re.search(r'\d', txt):
                    result["kudos"] = re.sub(r'\D', '', txt) or "0"
                    break
        except Exception:
            pass

    except Exception as e:
        print(f"    [상세 오류] {url} → {e}")

    return result


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 목록 페이지 파싱
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def parse_list_page(driver) -> list[dict]:
    """현재 열린 검색결과 페이지에서 게시글 기본 정보 추출"""
    items = []
    try:
        rows = driver.find_elements(By.CSS_SELECTOR, SEL_RESULT_ITEMS)
    except Exception:
        return items

    for row in rows:
        title = safe_text(row, SEL_TITLE)
        url   = safe_attr(row, SEL_TITLE, "href")
        if not title or not url:
            continue
        # 상대 URL 보정
        if url.startswith("/"):
            url = "https://r1.community.samsung.com" + url

        author  = safe_text(row, SEL_AUTHOR)
        board   = safe_text(row, SEL_BOARD)
        date_raw= safe_text(row, SEL_DATE_LIST)
        kudos   = safe_text(row, SEL_KUDOS_LIST) or "0"
        replies = safe_text(row, SEL_REPLIES_LIST) or "0"

        # 목록에서 날짜 먼저 확인 (빠른 필터링용)
        dt = parse_date(date_raw)

        items.append({
            "제목":      title,
            "작성자":    author,
            "게시판":    board,
            "작성일_raw": date_raw,
            "date_obj":  dt,
            "좋아요":    re.sub(r'\D', '', kudos) or "0",
            "댓글수":    re.sub(r'\D', '', replies) or "0",
            "URL":       url,
        })
    return items


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 메인 크롤링 로직
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def crawl() -> list[dict]:
    driver = build_driver()
    final  = []
    stop   = False
    page   = 1

    try:
        driver.get(SEARCH_URL)
        time.sleep(3)

        while not stop:
            print(f"\n[페이지 {page}] 로드 중...")

            # 목록 렌더링 대기
            try:
                wait_for(driver, SEL_RESULT_ITEMS, timeout=PAGE_TIMEOUT)
            except TimeoutException:
                print("  → 검색 결과 없음 또는 타임아웃. 종료.")
                break

            candidates = parse_list_page(driver)
            if not candidates:
                print("  → 게시글 없음. 종료.")
                break

            print(f"  → {len(candidates)}건 발견")

            for item in candidates:
                dt = item["date_obj"]

                # 날짜가 명확하고 2주 초과이면 이후 페이지도 스킵
                if dt is not None and dt < TWO_WEEKS_AGO:
                    print(f"  → '{item['제목'][:30]}' 날짜({dt.strftime('%Y-%m-%d')}) 2주 초과 → 수집 중단")
                    stop = True
                    break

                # 상세 페이지에서 본문 전체 + 정확한 날짜 + 좋아요 재수집
                print(f"  → 상세 수집: {item['제목'][:40]}")
                detail = fetch_detail(driver, item["URL"])

                # 상세에서 얻은 날짜가 있으면 덮어쓰기
                if detail["date"]:
                    item["date_obj"]  = detail["date"]
                    item["작성일_str"] = detail["date_str"]
                else:
                    item["작성일_str"] = item["작성일_raw"]

                # 상세 날짜로 재필터링
                if item["date_obj"] is not None and item["date_obj"] < TWO_WEEKS_AGO:
                    print(f"    ↳ 상세 날짜 {item['작성일_str']} → 2주 초과, 제외")
                    stop = True
                    break

                # 좋아요를 상세에서 얻었으면 덮어쓰기
                if detail["kudos"] and detail["kudos"] != "0":
                    item["좋아요"] = detail["kudos"]

                item["본문"] = detail["body"]
                final.append(item)

                # 목록 페이지로 복귀
                driver.back()
                time.sleep(2)

            if stop:
                break

            # 다음 페이지
            if MAX_PAGES and page >= MAX_PAGES:
                print(f"  → 최대 페이지({MAX_PAGES}) 도달. 종료.")
                break

            try:
                next_btn = driver.find_element(By.CSS_SELECTOR, SEL_NEXT_PAGE)
                next_btn.click()
                page += 1
                time.sleep(3)
            except NoSuchElementException:
                print("  → 다음 페이지 없음. 종료.")
                break

    finally:
        driver.quit()

    return final


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 엑셀 저장
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def save_excel(posts: list[dict], filepath: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "GPS 게시글 (2주이내)"

    # ── 스타일 ──
    thin   = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", start_color="1428A0")   # 삼성 블루
    c_align   = Alignment(horizontal="center", vertical="top",  wrap_text=True)
    l_align   = Alignment(horizontal="left",   vertical="top",  wrap_text=True)
    alt_fill  = PatternFill("solid", start_color="EEF2FB")

    # ── 헤더 ──
    headers = [
        ("번호",   6),
        ("제목",   45),
        ("작성자", 14),
        ("작성일", 20),
        ("게시판", 20),
        ("좋아요", 8),
        ("댓글수", 8),
        ("본문",   80),    # 본문 전체
        ("URL",    55),
    ]

    for ci, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = c_align
        cell.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    # ── 데이터 ──
    col_keys = ["번호","제목","작성자","작성일","게시판","좋아요","댓글수","본문","URL"]
    center_cols = {"번호","작성일","좋아요","댓글수"}

    for ri, post in enumerate(posts, 2):
        is_alt = (ri % 2 == 0)
        row_fill = alt_fill if is_alt else None

        for ci, key in enumerate(col_keys, 1):
            if key == "번호":
                val = ri - 1
            elif key == "작성일":
                val = post.get("작성일_str", post.get("작성일_raw", ""))
            else:
                val = post.get(key, "")

            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.font   = Font(name="Arial", size=10)
            if row_fill:
                cell.fill = row_fill

            if key == "URL" and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=10, color="1155CC", underline="single")

            cell.alignment = (
                c_align if key in center_cols else l_align
            )

        # 본문이 긴 경우 행 높이 자동 조절 (최대 400pt)
        body_len = len(str(post.get("본문", "")))
        row_h = min(400, max(30, body_len // 80 * 15 + 30))
        ws.row_dimensions[ri].height = row_h

    # ── 수집 정보 시트 ──
    ws2 = wb.create_sheet("수집 정보")
    info = [
        ("수집 일시",    datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("검색 URL",    SEARCH_URL),
        ("수집 기간",    f"최근 {DAYS_LIMIT}일 ({TWO_WEEKS_AGO.strftime('%Y-%m-%d')} ~ 오늘)"),
        ("수집 게시글 수", len(posts)),
        ("플랫폼",      "삼성멤버스 (Khoros 기반, Selenium + BeautifulSoup4 파싱)"),
    ]
    for r, (k, v) in enumerate(info, 1):
        c1 = ws2.cell(row=r, column=1, value=k)
        c1.font = Font(name="Arial", bold=True, size=10)
        c2 = ws2.cell(row=r, column=2, value=str(v))
        c2.font = Font(name="Arial", size=10)
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 80

    wb.save(filepath)
    print(f"\n✅ 저장 완료: {filepath}  (총 {len(posts)}건)")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 실행
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
if __name__ == "__main__":
    print("=" * 60)
    print("삼성멤버스 GPS 검색 크롤러")
    print(f"수집 기간: 최근 {DAYS_LIMIT}일 ({TWO_WEEKS_AGO.strftime('%Y-%m-%d')} 이후)")
    print("=" * 60)

    posts = crawl()

    if posts:
        save_excel(posts, OUTPUT_FILE)
    else:
        print("\n수집된 게시글이 없습니다.")
        print("가능한 원인:")
        print("  1. 최근 2주 이내 GPS 관련 게시글이 없음")
        print("  2. 삼성멤버스 사이트 구조 변경으로 셀렉터 불일치")
        print("  3. 봇 탐지로 인한 차단 (HEADLESS=False 로 재시도)")
